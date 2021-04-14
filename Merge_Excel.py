import openpyxl as xl
import pandas as pd


# opening the source excel file
path = '' #location of your file
filename = f'{path}'
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

# opening the destination excel file
path1 = '' #location of your file
filename1 = f'{path1}'
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

# copying the cell values from source
# excel file to destination excel file
for i in range(1, mr + 1):
    for j in range(1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row=i, column=j)

        # writing the read value to destination excel file
        ws2.cell(row=i, column=j).value = c.value

    # saving the destination excel file
wb2.save(str(filename1))

data = pd.read_excel(f'{path}')
data.drop_duplicates(subset = ["First Name"], keep = "first")
